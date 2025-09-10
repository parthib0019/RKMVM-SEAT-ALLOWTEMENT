import os
import io
import random as rnd
import pymysql
import openpyxl
from flask import Flask, request, redirect, flash, render_template, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

# ------------------ Flask Setup ------------------
app = Flask(__name__)
app.secret_key = "secret"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
OUTPUT_FOLDER = os.path.join(BASE_DIR, "output")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["OUTPUT_FOLDER"] = OUTPUT_FOLDER

# ------------------ Input Parsing ------------------
def parse_line(line: str):
    """
    Format: Date#RoomNumber!paperName@year%Subject^SubjectType
    Returns: (date_str, room_str, paper, year_str, subject, subject_type)
    """
    if not line or "#" not in line or "!" not in line or "@" not in line or "%" not in line or "^" not in line:
        raise ValueError(f"Bad line format: {line}")

    date_part, rest = line.split("#", 1)
    room_part, rest = rest.split("!", 1)
    paper_part, rest = rest.split("@", 1)
    year_part, rest = rest.split("%", 1)
    subject_part, subject_type_part = rest.split("^", 1)

    return (
        date_part.strip(),
        room_part.strip(),
        paper_part.strip(),
        year_part.strip(),
        subject_part.strip(),
        subject_type_part.strip().lower()
    )

# ------------------ Database Helper ------------------
def get_rolls_by_subject(year, subject, subject_type):
    
    
    print("fetching rolls")
    
    rolls = []
    conn = pymysql.connect(
        host="localhost", user="root", password="", database="ExamSeatAllowtment"
    )
    cursor = conn.cursor()

    query = f"SELECT student_data FROM StudentInfo WHERE Year LIKE '%{year}%'"
    cursor.execute(query)
    result = cursor.fetchone()
    if not result:
        cursor.close()
        conn.close()
        return rolls

    excel_blob = result[0]
    wb = openpyxl.load_workbook(io.BytesIO(excel_blob))
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    roll_idx = headers.index("Roll Number")
    honours_idx = headers.index("Honours")
    gen1_idx = headers.index("General1")
    gen2_idx = headers.index("General2")

    for row in ws.iter_rows(min_row=2, values_only=True):
        roll = row[roll_idx]
        honours = str(row[honours_idx] or "").lower()
        gen1 = str(row[gen1_idx] or "").lower()
        gen2 = str(row[gen2_idx] or "").lower()
        subj = subject.lower()

        if subject_type == "major":
            if subj in honours:
                rolls.append(roll)

        elif subject_type == "minor":
            year_num = int(year.split("-")[1])  # UG-1 → 1
            if year_num % 2 == 1:  # odd year → General1
                if subj in gen1:
                    rolls.append(roll)
            else:  # even year → General2
                if subj in gen2:
                    rolls.append(roll)

        elif subject_type == "general":
            year_num = int(year.split("-")[1])
            if subj in honours:
                rolls.append(roll)
            if year_num % 2 == 1 and subj in gen1:
                rolls.append(roll)
            if year_num % 2 == 0 and subj in gen2:
                rolls.append(roll)

    cursor.close()
    conn.close()
    return rolls

def get_room_info(room_id):
    try:
        # Connect to MySQL
        conn = pymysql.connect(
            host="localhost",       # or "127.0.0.1"
            user="root",            # default XAMPP MySQL user
            password="",            # set your password here if you have one
            database="ExamSeatAllowtment"   # <-- replace with your database name
        )
        cursor = conn.cursor()

        # Take input from terminal
        room_id = room_id.strip()

        # Query the database
        query = "SELECT RoomId, TotalCapacity, BenchPerCol FROM RoomInfo WHERE RoomId = %s"
        cursor.execute(query, (room_id,))

        result = cursor.fetchone()

        if result:
            # Bench arrangement
            bench_arrangement = []
            BenPerCols = result[2].split(",")
            BenPerCols.reverse()
            for i in BenPerCols:
                oneCol =[]
                for j in range(int(i)):
                    oneCol.append("e")
                bench_arrangement.append(oneCol)
                bench_arrangement.append(oneCol.copy())
            return bench_arrangement
        else:
            print(f"No room found with ID '{room_id}'")
        
        cursor.close()
        conn.close()

    except pymysql.Error as err:
        print(f"Error: {err}")

# ------------------ Allocation Helpers ------------------
def can_place(seat_matrix, c, r, paper, sep):
    separation = int(sep)
    for dc in range(-1*separation, separation):
        for dr in range(-1*separation, separation):
            if dc == 0 and dr == 0:
                continue
            cc = c + dc
            rr = r + dr
            if 0 <= cc < len(seat_matrix) and 0 <= rr < len(seat_matrix[cc]):
                neighbor = seat_matrix[cc][rr]
                #print(f"Neighbor at ({cc}, {rr}): {neighbor}")
                if neighbor != "e" and neighbor[1] == paper:
                    return False    
    return True

def allocate_seats(seat_matrix, rolls, paper, year, sep, subject):
    if not seat_matrix:
        return seat_matrix

    #rolls = list(rolls)  # make mutable copy

    for c in range(len(seat_matrix)):          # loop over columns
        for r in range(len(seat_matrix[c])):   # loop benches in column
            if not rolls:                      # stop if no rolls left
                seat_matrix.reverse()
                return seat_matrix

            if seat_matrix[c][r] == "e" and can_place(seat_matrix, c, r, paper, sep):
                roll = rolls.pop(0)            # assign one roll only
                seat_matrix[c][r] = (roll, paper, year, subject)
    return seat_matrix

# ------------------ PDF Export ------------------
def export_pdf(pdf_path, totalRooms):
    """
    totalRooms = {
        "Room15-2025-08-25": (seat_matrix, date),
        ...
    }
    """
    doc = SimpleDocTemplate(pdf_path, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # Constants for cell sizes
    SEAT_WIDTH = 80   # ~ 11 characters
    GUTTER_WIDTH = 15 # ~ 1 character
    ROW_HEIGHT = 25

    for room, (seat_matrix, date) in totalRooms.items():
        # --- Header ---
        room_display = room.split("_")[0]  # original room number
        header_text = (
            "<b>RAMAKRISHNA MISSION VIDYAMANDIRA</b><br/>"
            "Howrah, Belur: 711202<br/><br/>"
            f"<b>Date:</b> {date} &nbsp;&nbsp;&nbsp; <b>Room:</b> {room_display}<br/>"
        )
        elements.append(Paragraph(header_text, styles["Title"]))
        elements.append(Spacer(1, 12))

        # --- Build seat grid ---
        max_rows = max(len(col) for col in seat_matrix)
        data = []

        for r in range(max_rows):
            row_data = []
            for c in range(len(seat_matrix)):
                # Insert gutter after every 2 seat-columns
                if c > 0 and c % 2 == 0:
                    row_data.append("   ")

                if r < len(seat_matrix[c]):
                    seat = seat_matrix[c][r]
                    if seat == "e":  # empty seat with border
                        row_data.append("")  
                    elif seat is None:  # no seat at all
                        row_data.append(None)
                    else:  # filled seat
                        roll, paper, yr, subject = seat
                        row_data.append(f"{roll}\n{subject}-{yr}")
                else:
                    row_data.append(None)
            data.append(row_data)

        # --- Create custom colWidths ---
        num_cols = len(data[0])
        colWidths = []
        for c in range(num_cols):
            # If this column is gutter
            if all(row[c] == "   " or row[c] is None for row in data):
                colWidths.append(GUTTER_WIDTH)
            else:
                colWidths.append(SEAT_WIDTH)

        rowHeights = [ROW_HEIGHT for _ in range(len(data))]
        table = Table(data, colWidths=colWidths, rowHeights=rowHeights)

        # --- Styling ---
        style_commands = []

        for r, row in enumerate(data):
            for c, cell in enumerate(row):
                if cell is None:  # no seat → no border
                    style_commands.append(("BOX", (c, r), (c, r), 0, colors.white))
                elif cell == "   ":  # gutter → no border
                    style_commands.append(("BOX", (c, r), (c, r), 0, colors.white))
                else:  # filled seat or empty seat → border
                    style_commands.append(("GRID", (c, r), (c, r), 0.5, colors.black))

        # Merge gutters
        for c in range(num_cols):
            if all(row[c] == "   " for row in data):
                style_commands.append(("SPAN", (c, 0), (c, len(data)-1)))
                style_commands.append(("BOX", (c, 0), (c, len(data)-1), 0, colors.white))
                style_commands.append(("BACKGROUND", (c, 0), (c, len(data)-1), colors.white))

        style_commands.append(("ALIGN", (0, 0), (-1, -1), "CENTER"))
        style_commands.append(("VALIGN", (0, 0), (-1, -1), "MIDDLE"))
        table.setStyle(TableStyle(style_commands))

        elements.append(table)
        elements.append(PageBreak())

    doc.build(elements)

# ------------------ Routes ------------------
@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        SubjectDictionary = {"PHYSA": "Physics", "CHMA": "Chemistry", "MTMA": "Mathematics","ZOOA": "Zoology","HISA": "History", "ENGA": "English", "BNGA": "Bengali", "SNSA": "Sanskrit", "PHIL": "Philosophy", "COMS": "Computer Science", "ECOA": "Economics", "POLA": "Political Science"}
        totalRooms = {}
        file = request.files.get("file")
        if not file:
            flash("No file uploaded")
            return redirect(request.url)

        filepath = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename))
        file.save(filepath)

        with open(filepath, "r") as f:
            lines = f.read().splitlines()

        rnd.shuffle(lines)
        for line in lines:
            if not line.strip():
                continue
            date, room, paper, year, subject, subject_type = parse_line(line)
            print(line)                                                                    #########################
            rolls = get_rolls_by_subject(year, SubjectDictionary[subject], subject_type)
            print(rolls)                                                                    #########################
            room_key = f"{room}_{date.replace('/', '-')}"

            if room_key in totalRooms:
                seat_matrix = totalRooms[room_key][0]
            else:
                seat_matrix = get_room_info(room)
                print(seat_matrix)                                                           ########################
                if not seat_matrix:
                    continue
            seat_matrix = allocate_seats(seat_matrix, rolls, paper, year, 1, subject)
            print(seat_matrix)                                                              #######################33333333333
            totalRooms[room_key] = (seat_matrix, date)

        pdf_path = os.path.join(app.config['OUTPUT_FOLDER'], "All_Seating_Allotments.pdf")
        export_pdf(pdf_path, totalRooms)
        return render_template("pdf-viewer.html", pdf_files=["All_Seating_Allotments.pdf"])

    return render_template("index.html")

@app.route("/download/<filename>")
def download_file(filename):
    return send_from_directory(app.config['OUTPUT_FOLDER'], filename, as_attachment=True)

# ------------------ Run ------------------
if __name__ == "__main__":
    app.run(debug=True)
